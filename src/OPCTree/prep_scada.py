from __future__ import with_statement

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import settings

def write_conf(scada_config):
	is_xlsx = settings.SCADA_CONF_TEMPLATE_FILE.rsplit('.',1)[1] == 'xlsx'
	if is_xlsx:
		wb = load_workbook(settings.SCADA_CONF_TEMPLATE_FILE)
	else:
		raise Exception("Should be .xlsx, but is ." + settings.SCADA_CONF_TEMPLATE_FILE.rsplit('.',1)[1])
	
	for table_name, table in scada_config.items():
		if not '>>' in table_name: continue
		# try:
		ws = wb[table_name]

		for rowNbr in range(len(table)):
			row_values = tuple(table[rowNbr])
			for colNbr in range(len(row_values)):
				val = row_values[colNbr]
				cell = ws.cell(row=rowNbr+3, column=colNbr+1)
				cell.number_format = '@'
				try:
					float(val.replace(',','.'))
					is_number = True
				except ValueError:
					is_number = False
				except TypeError:
					cell.alignment = Alignment(horizontal='left')
					continue
				except AttributeError:
					cell.alignment = Alignment(horizontal='left')
					continue
				val = str(val)
				cell = ws.cell(row=rowNbr+3, column=colNbr+1, value=val)
				cell.alignment = Alignment(horizontal='right') if is_number else Alignment(horizontal='left')
		# except:
			# pass
	wb.save(settings.SCADA_CONF_OUTPUT_FILE)

"""
if __name__ == "__main__":
	import settings
	import you_lib_with_root_obj
	root = you_lib_with_root_obj.name_of_root_type('<Top Level>',description='')
	write_conf(root.scada_conf())
"""

